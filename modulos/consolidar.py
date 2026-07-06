"""
modulos/consolidar.py

Toma los 3 archivos descargados de ORS (POS) y los consolida en UN solo
reporte con formato profesional de 3 hojas:

  1. Resumen Ejecutivo      → totales del día + ventas por forma de pago
  2. Detalle de Checks      → todos los checks cerrados de ambos restaurantes
  3. Mapeo Simphony → Opera → solo los ROOM CHARGE que van al PMS

Entradas (en la carpeta pos del día):
  ORS_General_YYYY-MM-DD.xlsx       → totales de control (ventas, descuentos, etc.)
  ORS_Corcovado_YYYY-MM-DD.xlsx     → detalle de checks de Corcovado
  ORS_TerraKitchen_YYYY-MM-DD.xlsx  → detalle de checks de Terra Kitchen

Salida:
  Ventas_YYYY-MM-DD_FINAL.xlsx      → el reporte consolidado

Si todo sale bien, borra los 3 archivos originales y deja solo el final.
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# ── Paleta de colores (idéntica al formato del reporte modelo) ──────────────
AZUL_OSCURO   = "1F3864"   # títulos principales y totales
AZUL_MEDIO    = "2E5090"   # subtítulos / secciones
AZUL_CLARO    = "EEF2F9"   # filas alternas (zebra)
BLANCO        = "FFFFFF"
NARANJA       = "C55A11"   # encabezados de tabla y fila destacada

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

ARIAL = "Arial"


def _fill(color):
    return PatternFill("solid", start_color=color)


def _leer_detalle(path: Path, restaurante: str) -> pd.DataFrame:
    """Lee un archivo de detalle de ORS (Corcovado o Terra Kitchen)."""
    df = pd.read_excel(path, sheet_name="Reports", header=6)
    df = df[df["Check Number"].notna()].copy()
    df["Restaurante"] = restaurante
    return df


def _leer_totales_general(path: Path) -> dict:
    """Extrae los totales de control del archivo General."""
    gen = pd.read_excel(path, sheet_name="Reports", header=None)
    valores = {}
    for i in range(len(gen)):
        fila = gen.iloc[i].values
        if len(fila) > 2 and pd.notna(fila[1]) and isinstance(fila[1], str) and pd.notna(fila[2]):
            try:
                valores[fila[1]] = float(fila[2])
            except (ValueError, TypeError):
                pass
    return valores


def _hora_str(valor) -> str:
    """Convierte el campo de cierre a HH:MM."""
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")
    try:
        # serial de Excel → fracción del día
        frac = float(valor) % 1
        minutos = round(frac * 24 * 60)
        return f"{minutos // 60:02d}:{minutos % 60:02d}"
    except (ValueError, TypeError):
        return ""


def consolidar_pos(carpeta_pos: Path, fecha_reporte: datetime = None,
                   borrar_originales: bool = True) -> Path:
    """
    Consolida los 3 archivos ORS de carpeta_pos en un reporte final.
    Devuelve la ruta del archivo final generado.
    """
    carpeta_pos = Path(carpeta_pos)
    if fecha_reporte is None:
        fecha_reporte = datetime.now() - timedelta(days=1)
    fecha_str = fecha_reporte.strftime("%Y-%m-%d")
    fecha_titulo = f"{fecha_reporte.day} DE {MESES_ES[fecha_reporte.month]} {fecha_reporte.year}"

    f_general = carpeta_pos / f"ORS_General_{fecha_str}.xlsx"
    f_corco   = carpeta_pos / f"ORS_Corcovado_{fecha_str}.xlsx"
    f_terra   = carpeta_pos / f"ORS_TerraKitchen_{fecha_str}.xlsx"

    for f in (f_general, f_corco, f_terra):
        if not f.exists():
            raise FileNotFoundError(f"No se encontró {f.name} en {carpeta_pos}")

    # ── Cargar datos ────────────────────────────────────────────────────────
    corco = _leer_detalle(f_corco, "Corcovado")
    terra = _leer_detalle(f_terra, "Terra Kitchen")
    todos = pd.concat([corco, terra], ignore_index=True)
    totales = _leer_totales_general(f_general)

    total_general = todos["Payment Amount"].sum()

    # Resumen por forma de pago
    resumen = (todos.groupby(["Restaurante", "Tender Type"])["Payment Amount"]
               .sum().reset_index().sort_values(["Restaurante", "Tender Type"]))

    # Room charges → Opera
    room_charges = todos[todos["Tender Type"] == "ROOM CHARGE"].copy()
    total_rc = room_charges["Payment Amount"].sum()

    # ── Construir workbook ──────────────────────────────────────────────────
    wb = Workbook()
    _construir_resumen(wb.active, fecha_titulo, totales, resumen, total_general,
                       total_rc, len(todos))
    _construir_detalle(wb.create_sheet("Detalle de Checks"), fecha_titulo, todos, total_general)
    _construir_mapeo(wb.create_sheet("Mapeo Simphony → Opera"), fecha_titulo, room_charges, total_rc)

    salida = carpeta_pos / f"Ventas_{fecha_str}_FINAL.xlsx"
    wb.save(salida)
    logger.info(f"Reporte consolidado generado: {salida.name}")

    # ── Borrar originales ───────────────────────────────────────────────────
    if borrar_originales:
        for f in (f_general, f_corco, f_terra):
            try:
                f.unlink()
                logger.info(f"Borrado original: {f.name}")
            except OSError as e:
                logger.warning(f"No se pudo borrar {f.name}: {e}")

    return salida


def _construir_resumen(ws, fecha_titulo, totales, resumen, total_general,
                       total_rc, n_checks):
    ws.title = "Resumen Ejecutivo"
    for col, w in zip("ABCDEFG", [20, 32, 5, 5, 16, 14, 26]):
        ws.column_dimensions[col].width = w

    # Título
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"REPORTE DE VENTAS DIARIAS  ·  {fecha_titulo}"
    c.font = Font(ARIAL, 15, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "Terra Kitchen  &  Corcovado  —  Todos los Revenue Centers"
    c.font = Font(ARIAL, 10, color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)
    c.alignment = Alignment(horizontal="center")

    # Sección: Resumen general
    ws.merge_cells("A4:G4")
    c = ws["A4"]
    c.value = "  RESUMEN GENERAL DEL DÍA"
    c.font = Font(ARIAL, 11, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)

    bruto    = totales.get("Gross Sales Before Disc.", 0)
    desc     = totales.get("Discounts", 0)
    neto     = totales.get("Gross Sales After Disc.", 0)
    servicio = totales.get("Service Charges", 0)
    voids    = totales.get("Voids", 0)
    recibos  = totales.get("Receipts", 0)
    total_dia = neto + servicio

    filas_resumen = [
        ("Ventas Brutas (sin descuentos)", bruto, False),
        ("Descuentos", desc, False),
        ("Ventas Netas", neto, False),
        ("Cargos de Servicio", servicio, False),
        ("TOTAL VENTAS DEL DÍA", total_dia, True),
        ("Anulaciones (Voids)", -abs(voids) if voids else 0, False),
        ("Recibos / Cobrado", recibos, False),
    ]
    fila = 5
    for etiqueta, valor, total in filas_resumen:
        ws.merge_cells(f"A{fila}:E{fila}")
        ws.merge_cells(f"F{fila}:G{fila}")
        ca, cf = ws[f"A{fila}"], ws[f"F{fila}"]
        ca.value, cf.value = etiqueta, valor
        if total:
            fnt = Font(ARIAL, 10, bold=True, color=BLANCO)
            fl = _fill(AZUL_OSCURO)
        else:
            fnt = Font(ARIAL, 10, color="000000")
            fl = _fill(AZUL_CLARO if fila % 2 == 1 else BLANCO)
        ca.font = cf.font = fnt
        ca.fill = cf.fill = fl
        ca.alignment = Alignment(horizontal="left")
        cf.alignment = Alignment(horizontal="center")
        cf.number_format = "$#,##0.00"
        fila += 1

    # Línea de checks
    ws.merge_cells("A12:G12")
    c = ws["A12"]
    cerrados = int(totales.get("Checks Paid", 0)) if totales.get("Checks Paid", 0) else n_checks
    c.value = f"  Checks cerrados: {n_checks}   |   Total transacciones: {n_checks}"
    c.font = Font(ARIAL, 9, color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)

    # Sección: Ventas por forma de pago
    ws.merge_cells("A14:G14")
    c = ws["A14"]
    c.value = "  VENTAS POR FORMA DE PAGO"
    c.font = Font(ARIAL, 11, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)

    # Encabezados de tabla
    encabezados = ["Restaurante", "Forma de Pago", "", "", "Monto (USD)", "% del Total", "Destino"]
    for col, texto in zip("ABCDEFG", encabezados):
        c = ws[f"{col}15"]
        c.value = texto
        c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
        c.fill = _fill(NARANJA)
        c.alignment = Alignment(horizontal="center")

    fila = 16
    for _, r in resumen.iterrows():
        es_rc = r["Tender Type"] == "ROOM CHARGE"
        destino = "→ Opera (Room Charge)" if es_rc else "Interno / Paquete"
        bg = AZUL_CLARO if fila % 2 == 0 else BLANCO
        ws.merge_cells(f"B{fila}:D{fila}")
        valores = [
            ("A", r["Restaurante"], "left"),
            ("B", r["Tender Type"], "left"),
            ("E", r["Payment Amount"], "center"),
            ("F", r["Payment Amount"] / total_general, "center"),
            ("G", destino, "center"),
        ]
        for col, val, al in valores:
            c = ws[f"{col}{fila}"]
            c.value = val
            c.font = Font(ARIAL, 10, color="000000")
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal=al)
            if col == "E":
                c.number_format = "$#,##0.00"
            elif col == "F":
                c.number_format = "0.0%"
        fila += 1

    # Total general
    ws.merge_cells(f"A{fila}:D{fila}")
    c = ws[f"A{fila}"]
    c.value = "TOTAL GENERAL"
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    for col, val in [("E", total_general), ("F", 1.0)]:
        c = ws[f"{col}{fila}"]
        c.value = val
        c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
        c.fill = _fill(AZUL_OSCURO)
        c.alignment = Alignment(horizontal="center")
        c.number_format = "$#,##0.00" if col == "E" else "0.0%"
    ws[f"G{fila}"].fill = _fill(AZUL_OSCURO)
    fila += 1

    # Fila destacada: total a Opera
    ws.merge_cells(f"A{fila}:D{fila}")
    c = ws[f"A{fila}"]
    c.value = "▶  Total cargado a habitación enviado a Opera"
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(NARANJA)
    for col, val in [("E", total_rc), ("F", total_rc / total_general)]:
        c = ws[f"{col}{fila}"]
        c.value = val
        c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
        c.fill = _fill(NARANJA)
        c.alignment = Alignment(horizontal="center")
        c.number_format = "$#,##0.00" if col == "E" else "0.0%"
    ws[f"G{fila}"].fill = _fill(NARANJA)


def _construir_detalle(ws, fecha_titulo, todos, total_general):
    for col, w in zip("ABCDEFG", [18, 22, 12, 12, 32, 16, 16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"DETALLE DE CHECKS CERRADOS — {fecha_titulo}"
    c.font = Font(ARIAL, 13, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    encabezados = ["Restaurante", "Empleado", "# Check", "Hora Cierre",
                   "Forma de Pago", "Monto (USD)", "Subtotal Check"]
    for col, texto in zip("ABCDEFG", encabezados):
        c = ws[f"{col}2"]
        c.value = texto
        c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
        c.fill = _fill(NARANJA)
        c.alignment = Alignment(horizontal="center")

    fila = 3
    for _, r in todos.iterrows():
        bg = AZUL_CLARO if fila % 2 == 1 else BLANCO
        valores = [
            ("A", r["Restaurante"], "left", None),
            ("B", r["Transaction Employee"], "left", None),
            ("C", int(r["Check Number"]), "center", "0"),
            ("D", _hora_str(r["Check Closed Date and Time"]), "center", None),
            ("E", r["Tender Type"], "left", None),
            ("F", r["Payment Amount"], "center", "$#,##0.00"),
            ("G", r["Check Subtotal"], "center", "$#,##0.00"),
        ]
        for col, val, al, fmt in valores:
            c = ws[f"{col}{fila}"]
            c.value = val
            c.font = Font(ARIAL, 10, color="000000")
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal=al)
            if fmt:
                c.number_format = fmt
        fila += 1

    # Total
    ws.merge_cells(f"A{fila}:E{fila}")
    c = ws[f"A{fila}"]
    c.value = "TOTAL"
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    c = ws[f"F{fila}"]
    c.value = total_general
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    c.number_format = "$#,##0.00"
    ws[f"G{fila}"].fill = _fill(AZUL_OSCURO)


def _construir_mapeo(ws, fecha_titulo, room_charges, total_rc):
    for col, w in zip("ABCDEF", [18, 22, 12, 12, 14, 22]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"MAPEO SIMPHONY → OPERA  ·  ROOM CHARGES  ·  {fecha_titulo}"
    c.font = Font(ARIAL, 13, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Transacciones de cargo a habitación enviadas al PMS Opera"
    c.font = Font(ARIAL, 10, color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)
    c.alignment = Alignment(horizontal="center")

    encabezados = ["Restaurante", "Empleado", "# Check", "Hora Cierre",
                   "Habitación*", "Monto Cargado (USD)"]
    for col, texto in zip("ABCDEF", encabezados):
        c = ws[f"{col}3"]
        c.value = texto
        c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
        c.fill = _fill(NARANJA)
        c.alignment = Alignment(horizontal="center")

    fila = 4
    for _, r in room_charges.iterrows():
        bg = AZUL_CLARO if fila % 2 == 0 else BLANCO
        valores = [
            ("A", r["Restaurante"], "left", None),
            ("B", r["Transaction Employee"], "left", None),
            ("C", int(r["Check Number"]), "center", "0"),
            ("D", _hora_str(r["Check Closed Date and Time"]), "center", None),
            ("E", "—", "center", None),
            ("F", r["Payment Amount"], "center", "$#,##0.00"),
        ]
        for col, val, al, fmt in valores:
            c = ws[f"{col}{fila}"]
            c.value = val
            c.font = Font(ARIAL, 10, color="000000")
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal=al)
            if fmt:
                c.number_format = fmt
        fila += 1

    # Total
    ws.merge_cells(f"A{fila}:E{fila}")
    c = ws[f"A{fila}"]
    c.value = "TOTAL ENVIADO A OPERA"
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    c = ws[f"F{fila}"]
    c.value = total_rc
    c.font = Font(ARIAL, 10, bold=True, color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center")
    c.number_format = "$#,##0.00"
    fila += 1

    # Nota al pie
    ws.merge_cells(f"A{fila}:F{fila}")
    c = ws[f"A{fila}"]
    c.value = ("* Número de habitación disponible en el log de interfaz (.evt). "
               "El reporte de R&A no lo incluye por privacidad.")
    c.font = Font(ARIAL, 8, italic=True, color="666666")
    c.alignment = Alignment(horizontal="left")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    import sys
    carpeta = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    # Por defecto NO borra en modo prueba manual
    consolidar_pos(carpeta, borrar_originales=False)
